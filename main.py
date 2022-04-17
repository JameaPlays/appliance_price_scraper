import datetime
import time
import openpyxl.drawing.image
import urllib3
import io
import math
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, Border, PatternFill, Side, Alignment
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    from openpyxl.utils import column_index_from_string
from category import *


PRODUCT_BRAND_COL = 1
PRODUCT_MODEL_COL = 2
PRODUCT_DETAILS_COL = 3
PRODUCT_IMAGE_COL = 4
PRODUCT_IMAGE_COL_LETTER = 'D'
SENQ_PRICE_COL = 5
HN_PRICE_COL = 6
BHB_PRICE_COL = 7

categories = [fridge, washer_dryer, washing_machine, tv, soundbars, aircond, microwave_oven, cooking_hood, hob_stove]
headers = ['Brand', 'Model', 'Description', 'Image', 'SenQ', 'Harvey Norman', 'BHB', 'Min Price']
thin = Side(style=None, color='FF000000', border_style='thin')
http = urllib3.PoolManager()

wb = Workbook()
wb.remove(wb["Sheet"])

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver.maximize_window()


def check_duplicate(brand, model):
    duplicate = False
    # If product brand is included in category list
    if brand in category["brands"]:
        # Check for existing model number in worksheet
        for cell in ws['B']:
            # If model number already exists
            if cell.value == model:
                return cell


def senq_scrape():
    global excel_row_num
    # Scroll down 20 times
    for _ in range(20):
        driver.execute_script("window.scrollBy(0, document.body.scrollHeight)")
        time.sleep(1.5)

    # Get all products in product page
    all_products = driver.find_elements(By.CLASS_NAME, 'grid-item')
    for product in all_products:
        product_full_name = product.find_element(By.CLASS_NAME, 'MuiTypography-body1').text
        print(product_full_name)
        product_brand = product_full_name.split(' ')[0]
        product_model = product_full_name.split(' ')[-1]

        # If product brand is included in category list
        if product_brand in category["brands"]:

            # Check for existing model number
            existing = check_duplicate(product_brand, product_model)

            # If model number already exists, update price only
            if existing:
                try:
                    product_price = product.find_elements(By.CSS_SELECTOR, '.desc-item__price span')[1].text
                except IndexError:
                    product_price = product.find_element(By.CLASS_NAME, 'price_text').text
                chars = ['RM ', ',']
                for char in chars:
                    product_price = product_price.replace(char, '')
                try:
                    product_price = float(product_price)
                except ValueError:
                    pass
                ws.cell(row=existing.row, column=SENQ_PRICE_COL, value=product_price)
            else:
                # Product Brand
                ws.cell(row=excel_row_num, column=PRODUCT_BRAND_COL, value=product_brand)

                # Product Model number
                ws.cell(row=excel_row_num, column=PRODUCT_MODEL_COL, value=product_model)

                # Product Details
                product_details = ' '.join(product_full_name.split(' ')[1:-1])
                ws.cell(row=excel_row_num, column=PRODUCT_DETAILS_COL, value=product_details)

                # Product Image
                valid_img = False
                while not valid_img:
                    product_image_src = product.find_element(By.CLASS_NAME, 'img-bg-load').get_attribute('src')
                    if product_image_src[0:4] == "http":
                        valid_img = True
                # Get image from URL
                r = http.request('GET', product_image_src)
                image_file = io.BytesIO(r.data)
                # Insert image into excel cell
                img = openpyxl.drawing.image.Image(image_file)
                img.height = 210
                img.width = 290
                img.anchor = ws.cell(row=excel_row_num, column=PRODUCT_IMAGE_COL).coordinate
                ws.add_image(img)

                # Product Price
                try:
                    product_price = product.find_elements(By.CSS_SELECTOR, '.desc-item__price span')[1].text
                except IndexError:
                    product_price = product.find_element(By.CLASS_NAME, 'price_text').text
                chars = ['RM ', ',']
                for char in chars:
                    product_price = product_price.replace(char, '')
                try:
                    product_price = float(product_price)
                except ValueError:
                    pass
                ws.cell(row=excel_row_num, column=SENQ_PRICE_COL, value=product_price)

                # Autofit rows
                ws.row_dimensions[excel_row_num].height = 165

                # Next product (row)
                excel_row_num += 1


def hn_scrape():
    global excel_row_num

    # Number of pages based on total product count
    total_products = int(driver.find_element(By.CLASS_NAME, 'pagination-amount').text.split(' ')[0])
    no_of_pages = math.ceil(total_products / 20)

    # While next page is available
    for page in range(no_of_pages):

        # Get all products in product page
        all_products = driver.find_elements(By.CLASS_NAME, 'product-col')
        for product in all_products:
            product_full_name = product.find_element(By.CLASS_NAME, 'product-title').get_attribute('title')
            print(product_full_name)
            product_brand = product_full_name.split(' ')[0]
            product_model = product_full_name.split(' ')[1]
            chars = ['( ', ')']
            for char in chars:
                product_model = product_model.replace(char, '')

            # If product brand is included in category list
            if product_brand in category["brands"]:

                # Check for existing model number
                existing = check_duplicate(product_brand, product_model)

                # If model number already exists, update price only
                if existing:
                    product_price = product.find_elements(By.CLASS_NAME, 'price-num')[-1].text.replace(',', '')
                    try:
                        product_price = float(product_price)
                    except ValueError:
                        pass
                    ws.cell(row=existing.row, column=HN_PRICE_COL, value=product_price)
                else:
                    # Product Brand
                    ws.cell(row=excel_row_num, column=PRODUCT_BRAND_COL, value=product_brand)

                    # Product Model number
                    ws.cell(row=excel_row_num, column=PRODUCT_MODEL_COL, value=product_model)

                    # Product Details
                    product_details = ' '.join(product_full_name.split(' ')[2:])
                    ws.cell(row=excel_row_num, column=PRODUCT_DETAILS_COL, value=product_details)

                    # Product Image
                    product_image_src = product.find_element(By.TAG_NAME, 'img').get_attribute('src')
                    # Get image from URL
                    r = http.request('GET', product_image_src)
                    image_file = io.BytesIO(r.data)
                    # Insert image into excel cell
                    img = openpyxl.drawing.image.Image(image_file)
                    img.height = 210
                    img.width = 290
                    img.anchor = ws.cell(row=excel_row_num, column=PRODUCT_IMAGE_COL).coordinate
                    ws.add_image(img)

                    # Product Price
                    product_price = product.find_elements(By.CLASS_NAME, 'price-num')[-1].text.replace(',', '')
                    try:
                        product_price = float(product_price)
                    except ValueError:
                        pass
                    ws.cell(row=excel_row_num, column=HN_PRICE_COL, value=product_price)

                    # Autofit rows
                    ws.row_dimensions[excel_row_num].height = 165

                    # Next product (row)
                    excel_row_num += 1

        if page < (no_of_pages - 1):
            # Go to next page
            next_page_clicked = False
            while not next_page_clicked:
                try:
                    next_page = driver.find_elements(By.CSS_SELECTOR, '#pagination_contents li')[-1]
                    next_page.click()
                    time.sleep(2)
                    next_page_clicked = True
                # Close pop up if it appears
                except ElementClickInterceptedException:
                    pop_up = driver.find_element(By.CSS_SELECTOR, 'div.sp-fancybox-wrap iframe')
                    driver.switch_to.frame(pop_up)
                    driver.find_element(By.XPATH, '//*[@id="icon-close-button-1523407214825"]').click()
                    driver.switch_to.default_content()


def bhb_scrape():
    global excel_row_num

    # Number of pages based on total product count
    total_products = int(driver.find_element(By.CSS_SELECTOR, '.products-found strong').text)
    no_of_pages = math.ceil(total_products / 12)

    # While next page is available
    for page in range(no_of_pages):

        # Get all products in product page
        all_products = driver.find_elements(By.CLASS_NAME, 'product-inner')
        for product in all_products:
            product_full_name = product.find_element(By.CSS_SELECTOR, 'div.mf-product-content h2 a').text
            print(product_full_name)
            product_brand = product_full_name.split(' ')[0]
            product_model = product_full_name.split(' ')[1]

            # If product brand is included in category list
            if product_brand in category["brands"]:
                # Check for existing model number
                existing = check_duplicate(product_brand, product_model)

                # If model number already exists, update price only
                if existing:
                    product_price = product.find_elements(By.CSS_SELECTOR, '.price .woocommerce-Price-amount bdi')[
                        0].text
                    # if product_price == '':
                    #     product_price = product.find_elements(By.CSS_SELECTOR, '.price .woocommerce-Price-amount bdi')[
                    #         0].text
                    chars = ['RM', ',']
                    for char in chars:
                        product_price = product_price.replace(char, '')
                    try:
                        product_price = float(product_price)
                    except ValueError:
                        pass
                    ws.cell(row=existing.row, column=BHB_PRICE_COL, value=product_price)
                else:
                    # Product Brand
                    ws.cell(row=excel_row_num, column=PRODUCT_BRAND_COL, value=product_brand)

                    # Product Model number
                    ws.cell(row=excel_row_num, column=PRODUCT_MODEL_COL, value=product_model)

                    # Product Details
                    product_details = ' '.join(product_full_name.split(' ')[2:])
                    ws.cell(row=excel_row_num, column=PRODUCT_DETAILS_COL, value=product_details)

                    # Product Image
                    product_image_src = product.find_element(By.TAG_NAME, 'img').get_attribute('src')
                    # Get image from URL
                    r = http.request('GET', product_image_src)
                    image_file = io.BytesIO(r.data)
                    # Insert image into excel cell
                    img = openpyxl.drawing.image.Image(image_file)
                    img.height = 210
                    img.width = 290
                    img.anchor = ws.cell(row=excel_row_num, column=PRODUCT_IMAGE_COL).coordinate
                    ws.add_image(img)

                    # Product Price
                    product_price = product.find_elements(By.CSS_SELECTOR, '.price .woocommerce-Price-amount bdi')[1].text
                    if product_price == '':
                        product_price = product.find_elements(By.CSS_SELECTOR, '.price .woocommerce-Price-amount bdi')[0].text
                    chars = ['RM', ',']
                    for char in chars:
                        product_price = product_price.replace(char, '')
                    try:
                        product_price = float(product_price)
                    except ValueError:
                        pass
                    ws.cell(row=excel_row_num, column=BHB_PRICE_COL, value=product_price)

                    # Autofit rows
                    ws.row_dimensions[excel_row_num].height = 165

                    # Next product (row)
                    excel_row_num += 1

        if page < (no_of_pages - 1):
            # Go to next page
            # next_page_clicked = False
            # while not next_page_clicked:
            #     try:
            next_page = driver.find_element(By.CLASS_NAME, 'next')
            next_page.click()
            time.sleep(2)


# For each appliance category
for category in categories:
    ws = wb.create_sheet(category["name"])
    # Sheet Title
    ws.append([category["name"]])
    ws['A1'].font = Font(size=20, bold=True)
    # Sheet update time
    ws.append([f"Updated at: {datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')}"])
    # Table headers & styling
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws[get_column_letter(col) + '3'].font = Font(bold=True)
        ws[get_column_letter(col) + '3'].fill = PatternFill(fill_type='solid',
                                                            start_color='BDD7EE',
                                                            end_color='BDD7EE')

    excel_row_num = 4
    # For each online store for the category
    for shop, webpage_list in category["website"].items():
        # For each search link for the online store
        for webpage in webpage_list:
            driver.get(webpage)
            # SenQ Website
            if shop == 'senq':
                # pass
                senq_scrape()
            # Harvey Norman Website
            elif shop == "harvey_norman":
                # pass
                hn_scrape()
            # BHB Website
            elif shop == "bhb":
                # pass
                bhb_scrape()

    # Min Price between sellers
    for row in range(4, ws.max_row + 1):
        ws[get_column_letter(len(headers)) + str(row)] = f"=MIN({get_column_letter(SENQ_PRICE_COL) + str(row)}:" \
                                                         f"{get_column_letter(BHB_PRICE_COL) + str(row)})"

    # Autofit columns
    for row_cells in ws.columns:
        new_row_length = max(len(str(cell.value)) for cell in row_cells[3:])
        new_column_letter = (get_column_letter(row_cells[0].column))
        if new_row_length > 0:
            ws.column_dimensions[new_column_letter].width = new_row_length * 1.23
    # Set column width for image column
    ws.column_dimensions[get_column_letter(SENQ_PRICE_COL):get_column_letter(BHB_PRICE_COL)].width = 14.14
    ws.column_dimensions[PRODUCT_IMAGE_COL_LETTER].width = 42.14
    # Border for all cells in table
    for row in range(3, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws[get_column_letter(col) + str(row)].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws[get_column_letter(col) + str(row)].alignment = Alignment(vertical='center')

    wb.save(f'Electrical Appliance Price {datetime.date.today()}.xlsx')
